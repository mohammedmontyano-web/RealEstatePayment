import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta
import re
import uuid
from io import BytesIO

# دالة لتسمية الأقساط
def installment_name(n):
    names = ["", "الأول", "الثاني", "الثالث", "الرابع", "الخامس", "السادس", "السابع", "الثامن",
             "التاسع", "العاشر", "الحادي عشر", "الثاني عشر", "الثالث عشر", "الرابع عشر",
             "الخامس عشر", "السادس عشر", "السابع عشر", "الثامن عشر", "التاسع عشر", "العشرون",
             "الحادي والعشرون", "الثاني والعشرون", "الثالث والعشرون", "الرابع والعشرون",
             "الخامس والعشرون", "السادس والعشرون", "السابع والعشرون", "الثامن والعشرون",
             "التاسع والعشرون", "الثلاثون", "الحادي والثلاثين", "الثاني والثلاثون", "الثالث والثلاثون",
             "الرابع والثلاثون", "الخامس والثلاثون", "السادس والثلاثون", "السابع والثلاثون",
             "الثامن والثلاثون", "التاسع والثلاثون", "الأربعون", "الحادي والأربعون", "الثاني والأربعون",
             "الثالث والأربعون", "الرابع والأربعون", "الخامس والأربعون", "السادس والأربعون",
             "السابع والأربعون", "الثامن والأربعون", "التاسع والأربعون", "الخمسون",
             "الحادي والخمسون", "الثاني والخمسون", "الثالث والخمسون", "الرابع والخمسون",
             "الخامس والخمسون", "السادس والخمسون", "السابع والخمسون", "الثامن والخمسون",
             "التاسع والخمسون", "الستون", "الحادي والستون", "الثاني والستون", "الثالث والستون",
             "الرابع والستون", "الخامس والستون", "السادس والستون", "السابع والستون",
             "الثامن والستون", "التاسع والستون", "السبعون"]
    if 1 <= n <= 70:
        return f"القسط {names[n]}"
    return f"القسط {n}"

# دالة لتحويل الأرقام إلى نصوص عربية
def number_to_arabic_text(num):
    units = ["", "واحد", "اثنان", "ثلاثة", "أربعة", "خمسة", "ستة", "سبعة", "ثمانية", "تسعة"]
    teens = ["عشرة", "أحد عشر", "اثنا عشر", "ثلاثة عشر", "أربعة عشر", "خمسة عشر", "ستة عشر", "سبعة عشر", "ثمانية عشر", "تسعة عشر"]
    tens = ["", "", "عشرون", "ثلاثون", "أربعون", "خمسون", "ستون", "سبعون", "ثمانون", "تسعون"]
    hundreds = ["", "مائة", "مائتان", "ثلاثمائة", "أربعمائة", "خمسمائة", "ستمائة", "سبعمائة", "ثمانمائة", "تسعمائة"]
    groups = ["", ["ألف", "ألفان", "آلاف", "ألف"], "مليون", "مليار", "تريليون"]

    if num == 0:
        return "صفر جنيه مصري"
    
    result = ""
    num = int(num)
    i = 0
    
    while num > 0:
        group_value = num % 1000
        num //= 1000
        
        if group_value > 0:
            group_text = ""
            hundred = group_value // 100
            ten = (group_value % 100) // 10
            unit = group_value % 10
            
            if hundred > 0:
                group_text = hundreds[hundred]
            
            if ten > 0 or unit > 0:
                if hundred > 0:
                    group_text += " و"
                
                if ten == 0:
                    if unit > 0:
                        group_text += units[unit]
                elif ten == 1:
                    group_text += teens[unit]
                else:
                    if unit > 0:
                        group_text += f"{units[unit]} و{tens[ten]}"
                    else:
                        group_text += tens[ten]
            
            if group_text:
                if i == 0:
                    result = f"{group_text} {result}"
                elif i == 1:
                    if group_value == 1:
                        result = f"{groups[i][0]}{' و' + result if result else ''}"
                    elif group_value == 2:
                        result = f"{groups[i][1]}{' و' + result if result else ''}"
                    elif 3 <= group_value <= 10:
                        result = f"{group_text} {groups[i][2]}{' و' + result if result else ''}"
                    else:
                        result = f"{group_text} {groups[i][3]}{' و' + result if result else ''}"
                else:
                    result = f"{group_text} {groups[i]}{' و' + result if result else ''}"
        i += 1
    
    return f"{result.strip()} جنيه مصري لا غير"

# دالة Ceiling
def ceiling(number):
    return int(-(-number // 1))

# إعداد واجهة Streamlit
st.title("نظام جدول دفعات العقارات")

# إنشاء جلسة لحفظ البيانات
if "step" not in st.session_state:
    st.session_state.step = 1
    st.session_state.data = {
        "unit_name": "",
        "start_date": None,
        "period_type": "",
        "periods_per_year": 0,
        "period_months": 0,
        "total_installments": 0,
        "years": 0,
        "total_price": 0.0,
        "down_payment": 0.0,
        "delivery_payment": 0.0,
        "add_extra_payment": "",
        "extra_installments": [],
        "extra_amounts": [],
        "add_annual_payment": "",
        "annual_payment": 0.0,
        "add_fixed_amount": "",
        "fixed_installments": [],
        "fixed_amount": 0.0,
        "add_date_payment": "",
        "date_payments": [],
        "date_amounts": [],
        "total_remaining": 0.0,
        "modify_finish_total": "",
        "finish_total": 0.0,
        "finish_down": 0.0,
        "finish_installments": 0
    }
# إضافة مفتاح مؤقت للدفعات الإضافية
if "extra_payment_temp" not in st.session_state:
    st.session_state.extra_payment_temp = {"installment_num": 1, "amount": 0.0}

# دالة للتحقق من صحة التاريخ
def is_valid_date(date_str):
    try:
        date = datetime.strptime(date_str, "%d-%m-%Y")
        if 2000 <= date.year <= 2100:
            return True, date
        else:
            st.error("الرجاء إدخال تاريخ بين عام 2000 و2100!")
            return False, None
    except ValueError:
        st.error("صيغة التاريخ غير صحيحة! الرجاء إدخال التاريخ بصيغة يوم-شهر-سنة (مثل: 01-09-2025).")
        return False, None

# دالة لتنظيف اسم الوحدة
def clean_unit_name(unit_name):
    unit_name = re.sub(r'[:\\*/?"<>|]', '', unit_name)
    return unit_name[:25]

# الخطوات المتسلسلة
if st.session_state.step == 1:
    st.header("الخطوة 1: إدخال اسم الوحدة")
    unit_name = st.text_input("أدخل اسم الوحدة (مثل: شقة 101 أو فيلا A1):")
    if st.button("التالي"):
        if unit_name:
            st.session_state.data["unit_name"] = clean_unit_name(unit_name)
            st.session_state.step = 2
            st.rerun()
        else:
            st.error("الرجاء إدخال اسم الوحدة!")
    if st.button("إلغاء"):
        st.session_state.step = 1
        st.session_state.data = {}
        st.session_state.extra_payment_temp = {"installment_num": 1, "amount": 0.0}
        st.rerun()

elif st.session_state.step == 2:
    st.header("الخطوة 2: إدخال تاريخ بداية الأقساط")
    date_input = st.text_input("أدخل تاريخ بداية الأقساط (بصيغة يوم-شهر-سنة، مثل: 01-09-2025):")
    if st.button("التالي"):
        valid, date = is_valid_date(date_input)
        if valid:
            st.session_state.data["start_date"] = date
            st.session_state.step = 3
            st.rerun()
    if st.button("الرجوع"):
        st.session_state.step = 1
        st.rerun()
    if st.button("إلغاء"):
        st.session_state.step = 1
        st.session_state.data = {}
        st.session_state.extra_payment_temp = {"installment_num": 1, "amount": 0.0}
        st.rerun()

elif st.session_state.step == 3:
    st.header("الخطوة 3: اختيار نوع التقسيط")
    period_type = st.selectbox("اختر نوع التقسيط:", ["اختر...", "شهري", "ربع سنوي", "نصف سنوي"])
    if st.button("التالي"):
        if period_type != "اختر...":
            st.session_state.data["period_type"] = period_type
            if period_type == "شهري":
                st.session_state.data["periods_per_year"] = 12
                st.session_state.data["period_months"] = 1
            elif period_type == "ربع سنوي":
                st.session_state.data["periods_per_year"] = 4
                st.session_state.data["period_months"] = 3
            else:
                st.session_state.data["periods_per_year"] = 2
                st.session_state.data["period_months"] = 6
            st.session_state.step = 4
            st.rerun()
        else:
            st.error("الرجاء اختيار نوع التقسيط!")
    if st.button("الرجوع"):
        st.session_state.step = 2
        st.rerun()
    if st.button("إلغاء"):
        st.session_state.step = 1
        st.session_state.data = {}
        st.session_state.extra_payment_temp = {"installment_num": 1, "amount": 0.0}
        st.rerun()

elif st.session_state.step == 4:
    st.header("الخطوة 4: إدخال عدد الشهور أو السنوات")
    if st.session_state.data["period_type"] == "شهري":
        months = st.number_input("أدخل عدد الشهور للتقسيط (من 1 إلى 120):", min_value=1, max_value=120, step=1)
        if st.button("التالي"):
            st.session_state.data["total_installments"] = months
            st.session_state.data["years"] = ceiling(months / 12)
            st.session_state.step = 5
            st.rerun()
    else:
        years = st.number_input("أدخل عدد سنوات التقسيط:", min_value=1, step=1)
        if st.button("التالي"):
            st.session_state.data["years"] = years
            st.session_state.data["total_installments"] = years * st.session_state.data["periods_per_year"]
            st.session_state.step = 5
            st.rerun()
    if st.button("الرجوع"):
        st.session_state.step = 3
        st.rerun()
    if st.button("إلغاء"):
        st.session_state.step = 1
        st.session_state.data = {}
        st.session_state.extra_payment_temp = {"installment_num": 1, "amount": 0.0}
        st.rerun()

elif st.session_state.step == 5:
    st.header("الخطوة 5: إدخال المبلغ الإجمالي للعقار")
    total_price = st.number_input("أدخل المبلغ الإجمالي للعقار (مثل: 1000000):", min_value=0.0, step=1000.0)
    if st.button("التالي"):
        if total_price > 0:
            st.session_state.data["total_price"] = total_price
            st.session_state.step = 6
            st.rerun()
        else:
            st.error("المبلغ الإجمالي يجب أن يكون أكبر من صفر!")
    if st.button("الرجوع"):
        st.session_state.step = 4
        st.rerun()
    if st.button("إلغاء"):
        st.session_state.step = 1
        st.session_state.data = {}
        st.session_state.extra_payment_temp = {"installment_num": 1, "amount": 0.0}
        st.rerun()

elif st.session_state.step == 6:
    st.header("الخطوة 6: إدخال المقدم الكلي")
    down_payment = st.number_input(f"أدخل مبلغ المقدم الكلي (من 0 إلى {st.session_state.data['total_price']}):", min_value=0.0, max_value=st.session_state.data["total_price"], step=1000.0)
    if st.button("التالي"):
        st.session_state.data["down_payment"] = down_payment
        st.session_state.step = 7
        st.rerun()
    if st.button("الرجوع"):
        st.session_state.step = 5
        st.rerun()
    if st.button("إلغاء"):
        st.session_state.step = 1
        st.session_state.data = {}
        st.session_state.extra_payment_temp = {"installment_num": 1, "amount": 0.0}
        st.rerun()

elif st.session_state.step == 7:
    st.header("الخطوة 7: إدخال دفعة الاستلام")
    max_delivery = st.session_state.data["total_price"] - st.session_state.data["down_payment"]
    delivery_payment = st.number_input(f"أدخل مبلغ دفعة الاستلام (من 0 إلى {max_delivery}):", min_value=0.0, max_value=max_delivery, step=1000.0)
    if st.button("التالي"):
        st.session_state.data["delivery_payment"] = delivery_payment
        st.session_state.step = 8
        st.rerun()
    if st.button("الرجوع"):
        st.session_state.step = 6
        st.rerun()
    if st.button("إلغاء"):
        st.session_state.step = 1
        st.session_state.data = {}
        st.session_state.extra_payment_temp = {"installment_num": 1, "amount": 0.0}
        st.rerun()

elif st.session_state.step == 8:
    st.header("الخطوة 8: إضافة دفعات إضافية")
    add_extra = st.selectbox("هل تريد إضافة دفعات إضافية في أقساط محددة؟", ["اختر...", "نعم", "لا"], key="add_extra_select")

    if add_extra == "نعم":
        st.subheader("إدخال دفعة إضافية")
        # عرض الحقول في حاوية منفصلة
        with st.container():
            col1, col2, col3 = st.columns([2, 2, 1])
            with col1:
                installment_num = st.number_input("أدخل رقم القسط:", min_value=1, max_value=st.session_state.data["total_installments"], step=1, key=f"installment_num_{st.session_state.get('extra_payment_key', 0)}")
            with col2:
                amount = st.number_input("أدخل مبلغ الدفعة الإضافية:", min_value=0.0, step=1000.0, key=f"amount_{st.session_state.get('extra_payment_key', 0)}")
            with col3:
                if st.button("تأكيد الدفعة", key=f"confirm_extra_{st.session_state.get('extra_payment_key', 0)}"):
                    if installment_num not in st.session_state.data["extra_installments"]:
                        st.session_state.data["extra_installments"].append(installment_num)
                        st.session_state.data["extra_amounts"].append(amount)
                        st.session_state.extra_payment_temp = {"installment_num": 1, "amount": 0.0}  # إعادة تعيين الحقول
                        st.session_state["extra_payment_key"] = st.session_state.get("extra_payment_key", 0) + 1  # تحديث المفتاح
                        st.success(f"تم إضافة دفعة إضافية للقسط {installment_num}")
                    else:
                        st.error("رقم القسط مكرر!")
        
        # عرض الدفعات الإضافية الحالية
        if st.session_state.data["extra_installments"]:
            st.subheader("الدفعات الإضافية المُدخلة")
            extra_data = pd.DataFrame({
                "رقم القسط": st.session_state.data["extra_installments"],
                "المبلغ": st.session_state.data["extra_amounts"]
            })
            st.table(extra_data)
            # إضافة زر إزالة لكل دفعة
            for i, (inst, amt) in enumerate(zip(st.session_state.data["extra_installments"], st.session_state.data["extra_amounts"])):
                if st.button(f"إزالة القسط {inst}", key=f"remove_extra_{i}"):
                    st.session_state.data["extra_installments"].pop(i)
                    st.session_state.data["extra_amounts"].pop(i)
                    st.rerun()

    if st.button("التالي"):
        if add_extra != "اختر...":
            st.session_state.data["add_extra_payment"] = "Y" if add_extra == "نعم" else "N"
            st.session_state.step = 9
            st.rerun()
        else:
            st.error("الرجاء اختيار نعم أو لا!")
    if st.button("الرجوع"):
        st.session_state.step = 7
        st.rerun()
    if st.button("إلغاء"):
        st.session_state.step = 1
        st.session_state.data = {}
        st.session_state.extra_payment_temp = {"installment_num": 1, "amount": 0.0}
        st.rerun()

elif st.session_state.step == 9:
    st.header("الخطوة 9: إضافة دفعات سنوية")
    annual_installments = [i * st.session_state.data["periods_per_year"] for i in range(1, st.session_state.data["years"] + 1)]
    st.write(f"أرقام الأقساط السنوية: {', '.join(map(str, annual_installments))}")
    add_annual = st.selectbox("هل تريد إضافة مبلغ إضافي في نهاية كل سنة؟", ["اختر...", "نعم", "لا"])
    if add_annual == "نعم":
        annual_payment = st.number_input("أدخل مبلغ الدفعة السنوية:", min_value=0.0, step=1000.0)
    if st.button("التالي"):
        if add_annual != "اختر...":
            st.session_state.data["add_annual_payment"] = "Y" if add_annual == "نعم" else "N"
            st.session_state.data["annual_payment"] = annual_payment if add_annual == "نعم" else 0.0
            st.session_state.step = 10
            st.rerun()
        else:
            st.error("الرجاء اختيار نعم أو لا!")
    if st.button("الرجوع"):
        st.session_state.step = 8
        st.rerun()
    if st.button("إلغاء"):
        st.session_state.step = 1
        st.session_state.data = {}
        st.session_state.extra_payment_temp = {"installment_num": 1, "amount": 0.0}
        st.rerun()

elif st.session_state.step == 10:
    st.header("الخطوة 10: إضافة دفعات ثابتة")
    add_fixed = st.selectbox("هل هناك دفعات ثابتة في أقساط محددة؟", ["اختر...", "نعم", "لا"])
    if add_fixed == "نعم":
        fixed_installments_input = st.text_input("أدخل أرقام الأقساط الثابتة (مفصولة بفواصل، مثل: 1,3,5):")
        fixed_amount = st.number_input("أدخل المبلغ الثابت:", min_value=0.0, step=1000.0)
        if st.button("تأكيد الأقساط الثابتة"):
            try:
                fixed_installments = [int(x) for x in fixed_installments_input.split(",")]
                if all(1 <= x <= st.session_state.data["total_installments"] for x in fixed_installments):
                    if len(fixed_installments) == len(set(fixed_installments)):
                        st.session_state.data["fixed_installments"] = fixed_installments
                        st.session_state.data["fixed_amount"] = fixed_amount
                        st.success("تم إضافة الأقساط الثابتة")
                    else:
                        st.error("لا يمكن تكرار أرقام الأقساط!")
                else:
                    st.error(f"الأقساط يجب أن تكون بين 1 و{st.session_state.data['total_installments']}")
            except ValueError:
                st.error("إدخال غير صحيح! الرجاء إدخال أرقام مفصولة بفواصل.")
    if st.button("التالي"):
        if add_fixed != "اختر...":
            st.session_state.data["add_fixed_amount"] = "Y" if add_fixed == "نعم" else "N"
            if add_fixed == "لا":
                st.session_state.data["fixed_installments"] = []
                st.session_state.data["fixed_amount"] = 0.0
            st.session_state.step = 11
            st.rerun()
        else:
            st.error("الرجاء اختيار نعم أو لا!")
    if st.button("الرجوع"):
        st.session_state.step = 9
        st.rerun()
    if st.button("إلغاء"):
        st.session_state.step = 1
        st.session_state.data = {}
        st.session_state.extra_payment_temp = {"installment_num": 1, "amount": 0.0}
        st.rerun()

elif st.session_state.step == 11:
    st.header("الخطوة 11: إضافة دفعات بتواريخ محددة")
    add_date_payment = st.selectbox("هل تريد إضافة دفعات في تواريخ محددة؟", ["اختر...", "نعم", "لا"])
    if add_date_payment == "نعم":
        payment_type = st.selectbox("اختر النوع:", ["اختر...", "تاريخ واحد", "تكرار سنوي"])
        date_str = st.text_input("أدخل التاريخ (dd-mm-yyyy، مثل: 04-05-2025):")
        amount = st.number_input("أدخل المبلغ:", min_value=0.0, step=1000.0)
        if st.button("تأكيد الدفعة"):
            valid, date = is_valid_date(date_str)
            if valid and payment_type != "اختر...":
                if payment_type == "تاريخ واحد":
                    st.session_state.data["date_payments"].append(date)
                    st.session_state.data["date_amounts"].append(amount)
                else:
                    for y in range(st.session_state.data["years"]):
                        this_date = date + timedelta(days=365 * y)
                        st.session_state.data["date_payments"].append(this_date)
                        st.session_state.data["date_amounts"].append(amount)
                st.success("تم إضافة الدفعة")
    if st.button("التالي"):
        if add_date_payment != "اختر...":
            st.session_state.data["add_date_payment"] = "Y" if add_date_payment == "نعم" else "N"
            st.session_state.step = 12
            st.rerun()
        else:
            st.error("الرجاء اختيار نعم أو لا!")
    if st.button("الرجوع"):
        st.session_state.step = 10
        st.rerun()
    if st.button("إلغاء"):
        st.session_state.step = 1
        st.session_state.data = {}
        st.session_state.extra_payment_temp = {"installment_num": 1, "amount": 0.0}
        st.rerun()

elif st.session_state.step == 12:
    # حساب المبلغ المتبقي
    total_extra_amount = sum(st.session_state.data["extra_amounts"])
    total_annual_amount = st.session_state.data["annual_payment"] * st.session_state.data["years"]
    total_fixed_amount = st.session_state.data["fixed_amount"] * len(st.session_state.data["fixed_installments"])
    total_date_amount = sum(st.session_state.data["date_amounts"])
    total_remaining = (st.session_state.data["total_price"] - st.session_state.data["down_payment"] -
                      st.session_state.data["delivery_payment"] - total_fixed_amount - total_extra_amount -
                      total_annual_amount - total_date_amount)
    
    st.header("الخطوة 12: تأكيد المبلغ المتبقي")
    st.write(f"المبلغ المتبقي بعد خصم المقدمة ({st.session_state.data['down_payment']})، دفعة الاستلام "
             f"({st.session_state.data['delivery_payment']})، الدفعات الإضافية ({total_extra_amount + total_annual_amount})، "
             f"الدفعات الثابتة ({total_fixed_amount})، والدفعات بتواريخ محددة ({total_date_amount}) هو: {total_remaining}")
    confirm = st.selectbox("هل تريد توزيعه كما هو أم تعديله؟", ["اختر...", "توزيع كما هو", "تعديل"])
    if confirm == "تعديل":
        new_remaining = st.number_input("أدخل القيمة الجديدة للمتبقي:", min_value=0.0, step=1000.0)
    if st.button("التالي"):
        if confirm != "اختر...":
            st.session_state.data["total_remaining"] = new_remaining if confirm == "تعديل" else total_remaining
            if st.session_state.data["total_remaining"] >= 0:
                st.session_state.step = 13
                st.rerun()
            else:
                st.error("المبلغ المتبقي يجب أن يكون أكبر من أو يساوي صفر!")
        else:
            st.error("الرجاء اختيار خيار!")
    if st.button("الرجوع"):
        st.session_state.step = 11
        st.rerun()
    if st.button("إلغاء"):
        st.session_state.step = 1
        st.session_state.data = {}
        st.session_state.extra_payment_temp = {"installment_num": 1, "amount": 0.0}
        st.rerun()

elif st.session_state.step == 13:
    st.header("الخطوة 13: تحديد إجمالي التشطيب")
    modify_finish = st.selectbox("هل تريد تحديد إجمالي مبلغ التشطيب يدويًا؟", ["اختر...", "نعم", "لا"])
    if modify_finish == "نعم":
        suggested_finish = st.session_state.data["total_price"] * 0.4
        finish_total = st.number_input(f"أدخل إجمالي مبلغ التشطيب (المقترح: {suggested_finish}):", min_value=0.0, max_value=st.session_state.data["total_price"], step=1000.0)
    else:
        finish_total = st.session_state.data["total_price"] * 0.4
    if st.button("التالي"):
        if modify_finish != "اختر...":
            st.session_state.data["modify_finish_total"] = "Y" if modify_finish == "نعم" else "N"
            st.session_state.data["finish_total"] = finish_total
            st.session_state.step = 14
            st.rerun()
        else:
            st.error("الرجاء اختيار نعم أو لا!")
    if st.button("الرجوع"):
        st.session_state.step = 12
        st.rerun()
    if st.button("إلغاء"):
        st.session_state.step = 1
        st.session_state.data = {}
        st.session_state.extra_payment_temp = {"installment_num": 1, "amount": 0.0}
        st.rerun()

elif st.session_state.step == 14:
    st.header("الخطوة 14: إدخال مقدمة التشطيب")
    max_finish_down = min(st.session_state.data["finish_total"], st.session_state.data["down_payment"])
    finish_down = st.number_input(f"أدخل مبلغ مقدمة التشطيب (من 0 إلى {max_finish_down}):", min_value=0.0, max_value=max_finish_down, step=1000.0)
    if st.button("التالي"):
        st.session_state.data["finish_down"] = finish_down
        st.session_state.step = 15
        st.rerun()
    if st.button("الرجوع"):
        st.session_state.step = 13
        st.rerun()
    if st.button("إلغاء"):
        st.session_state.step = 1
        st.session_state.data = {}
        st.session_state.extra_payment_temp = {"installment_num": 1, "amount": 0.0}
        st.rerun()

elif st.session_state.step == 15:
    st.header("الخطوة 15: إدخال عدد أقساط التشطيب")
    max_finish_install = min(16, st.session_state.data["total_installments"])
    finish_installments = st.number_input("أدخل عدد أقساط التشطيب (من 0 إلى 16):", min_value=0, max_value=max_finish_install, step=1)
    if st.button("إنشاء الملف"):
        if (st.session_state.data["finish_total"] == 0 and finish_installments == 0) or \
           (st.session_state.data["finish_total"] > 0 and finish_installments > 0):
            st.session_state.data["finish_installments"] = finish_installments
            st.session_state.step = 16
            st.rerun()
        else:
            st.error("إذا كان إجمالي التشطيب أكبر من صفر، يجب أن تكون الأقساط أكبر من صفر، وإذا كان صفرًا، يجب أن تكون الأقساط صفرًا!")
    if st.button("الرجوع"):
        st.session_state.step = 14
        st.rerun()
    if st.button("إلغاء"):
        st.session_state.step = 1
        st.session_state.data = {}
        st.session_state.extra_payment_temp = {"installment_num": 1, "amount": 0.0}
        st.rerun()

elif st.session_state.step == 16:
    st.header("إنشاء جدول الدفعات")
    
    # استخراج البيانات
    data = st.session_state.data
    unit_name = data["unit_name"]
    start_date = data["start_date"]
    period_type = data["period_type"]
    periods_per_year = data["periods_per_year"]
    period_months = data["period_months"]
    total_installments = data["total_installments"]
    years = data["years"]
    total_price = data["total_price"]
    down_payment = data["down_payment"]
    delivery_payment = data["delivery_payment"]
    extra_installments = data["extra_installments"]
    extra_amounts = data["extra_amounts"]
    annual_payment = data["annual_payment"]
    fixed_installments = data["fixed_installments"]
    fixed_amount = data["fixed_amount"]
    date_payments = data["date_payments"]
    date_amounts = data["date_amounts"]
    total_remaining = data["total_remaining"]
    finish_total = data["finish_total"]
    finish_down = data["finish_down"]
    finish_installments = data["finish_installments"]
    
    # التحقق من عدم التداخل
    annual_installments = [i * periods_per_year for i in range(1, years + 1)]
    for i in annual_installments:
        if i in fixed_installments or i in extra_installments:
            st.error(f"القسط {i} مدرج كدفعة ثابتة أو إضافية! لا يمكن التداخل.")
            st.button("الرجوع", on_click=lambda: setattr(st.session_state, "step", 11))
            st.button("إلغاء", on_click=lambda: [setattr(st.session_state, "step", 1), st.session_state.data.clear(), setattr(st.session_state, "extra_payment_temp", {"installment_num": 1, "amount": 0.0})])
            st.stop()
    for i in fixed_installments:
        if i in extra_installments:
            st.error(f"القسط {i} مدرج كدفعة ثابتة وإضافية! لا يمكن التداخل.")
            st.button("الرجوع", on_click=lambda: setattr(st.session_state, "step", 10))
            st.button("إلغاء", on_click=lambda: [setattr(st.session_state, "step", 1), st.session_state.data.clear(), setattr(st.session_state, "extra_payment_temp", {"installment_num": 1, "amount": 0.0})])
            st.stop()
    
    # حساب المبالغ
    unit_down = down_payment - finish_down
    unit_total = total_price - finish_total
    finish_remaining = finish_total - finish_down
    unit_remaining = unit_total - unit_down - delivery_payment
    
    non_fixed_installments = total_installments - len(fixed_installments)
    if non_fixed_installments <= 0 and total_remaining > 0:
        st.error(f"جميع الأقساط ثابتة! يجب أن يكون هناك أقساط عادية لتوزيع المبلغ المتبقي ({total_remaining}).")
        st.button("الرجوع", on_click=lambda: setattr(st.session_state, "step", 10))
        st.stop()
    
    property_install = int(total_remaining / non_fixed_installments / 1000) * 1000 if non_fixed_installments > 0 else 0
    property_rem = total_remaining - (property_install * non_fixed_installments)
    if property_rem < 0:
        property_install = int((total_remaining / non_fixed_installments) / 1000) * 1000 - 1000
        property_rem = total_remaining - (property_install * non_fixed_installments)
        if property_rem < 0:
            property_rem = 0
    
    # جمع التواريخ
    all_dates = []
    all_install_nums = []
    all_types = []
    for i in range(1, total_installments + 1):
        temp_date = start_date + timedelta(days=30 * period_months * (i - 1))
        all_dates.append(temp_date)
        all_install_nums.append(i)
        all_types.append("regular")
    
    for j, date in enumerate(date_payments):
        if date not in all_dates:
            all_dates.append(date)
            all_install_nums.append(0)
            all_types.append("date_only")
    
    # ترتيب التواريخ
    sorted_indices = sorted(range(len(all_dates)), key=lambda k: all_dates[k])
    sorted_dates = [all_dates[i] for i in sorted_indices]
    sorted_install_nums = [all_install_nums[i] for i in sorted_indices]
    sorted_types = [all_types[i] for i in sorted_indices]
    
    # حساب تاريخ نهاية التشطيب
    finish_end_date = start_date + timedelta(days=30 * period_months * (finish_installments - 1)) if finish_installments > 0 else start_date
    
    # إعداد ملف Excel
    wb = Workbook()
    ws_unit = wb.create_sheet(title=f"{unit_name} - وحدة")
    ws_finish = wb.create_sheet(title=f"{unit_name} - تشطيب")
    ws_combined = wb.create_sheet(title=f"{unit_name} - TOTALLY")
    wb.remove(wb["Sheet"])
    
    # تنسيق الخلايا
    blue_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    cyan_fill = PatternFill(start_color="0099FF", end_color="0099FF", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    purple_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    calibri_font = Font(name="Calibri")
    
    # إعداد جدول الوحدة
    ws_unit["B1"] = f"جدول أقساط وحدة {unit_name}"
    ws_unit.merge_cells("B1:E1")
    for col in range(1, 6):
        ws_unit.cell(1, col).fill = blue_fill
        ws_unit.cell(1, col).font = white_font
        ws_unit.cell(1, col).alignment = Alignment(horizontal="center", vertical="center")
        ws_unit.cell(1, col).border = Border(bottom=Side(style="medium"))
    
    ws_unit["B2"] = "المقدمة"
    ws_unit["D2"] = unit_down
    ws_unit["E2"] = number_to_arabic_text(unit_down)
    for col in range(1, 6):
        ws_unit.cell(2, col).fill = yellow_fill
        ws_unit.cell(2, col).font = bold_font
        ws_unit.cell(2, col).alignment = Alignment(horizontal="right")
        ws_unit.cell(2, col).border = Border(bottom=Side(style="medium"))
    
    ws_unit["A3"] = "م"
    ws_unit["B3"] = "القسط"
    ws_unit["C3"] = "التاريخ"
    ws_unit["D3"] = "القسط"
    ws_unit["E3"] = "القسط (بالحروف)"
    for col in range(1, 6):
        ws_unit.cell(3, col).fill = cyan_fill
        ws_unit.cell(3, col).font = white_font
        ws_unit.cell(3, col).alignment = Alignment(horizontal="center")
        ws_unit.cell(3, col).border = Border(bottom=Side(style="medium"))
    
    # إعداد جدول التشطيب
    ws_finish["B1"] = f"جدول أقساط تشطيب {unit_name}"
    ws_finish.merge_cells("B1:E1")
    for col in range(1, 6):
        ws_finish.cell(1, col).fill = blue_fill
        ws_finish.cell(1, col).font = white_font
        ws_finish.cell(1, col).alignment = Alignment(horizontal="center", vertical="center")
        ws_finish.cell(1, col).border = Border(bottom=Side(style="medium"))
    
    ws_finish["B2"] = "مقدمة التشطيب"
    ws_finish["D2"] = finish_down
    ws_finish["E2"] = number_to_arabic_text(finish_down)
    for col in range(1, 6):
        ws_finish.cell(2, col).fill = yellow_fill
        ws_finish.cell(2, col).font = bold_font
        ws_finish.cell(2, col).alignment = Alignment(horizontal="right")
        ws_finish.cell(2, col).border = Border(bottom=Side(style="medium"))
    
    ws_finish["A3"] = "م"
    ws_finish["B3"] = "القسط"
    ws_finish["C3"] = "التاريخ"
    ws_finish["D3"] = "القسط"
    ws_finish["E3"] = "القسط (بالحروف)"
    for col in range(1, 6):
        ws_finish.cell(3, col).fill = cyan_fill
        ws_finish.cell(3, col).font = white_font
        ws_finish.cell(3, col).alignment = Alignment(horizontal="center")
        ws_finish.cell(3, col).border = Border(bottom=Side(style="medium"))
    
    # إعداد الجدول المجمع
    ws_combined["B1"] = f"جدول أقساط مجمع {unit_name}"
    ws_combined.merge_cells("B1:G1")
    for col in range(1, 8):
        ws_combined.cell(1, col).fill = blue_fill
        ws_combined.cell(1, col).font = white_font
        ws_combined.cell(1, col).alignment = Alignment(horizontal="center", vertical="center")
        ws_combined.cell(1, col).border = Border(bottom=Side(style="medium"))
    
    ws_combined["B2"] = "المقدمة الكلية"
    ws_combined["D2"] = down_payment
    ws_combined["E2"] = number_to_arabic_text(down_payment)
    for col in range(1, 8):
        ws_combined.cell(2, col).fill = yellow_fill
        ws_combined.cell(2, col).font = bold_font
        ws_combined.cell(2, col).alignment = Alignment(horizontal="right")
        ws_combined.cell(2, col).border = Border(bottom=Side(style="medium"))
    
    ws_combined["A3"] = "م"
    ws_combined["B3"] = "التاريخ"
    ws_combined["C3"] = "القسط"
    ws_combined["D3"] = "القسط الكلي"
    ws_combined["E3"] = "القسط (بالحروف)"
    ws_combined["F3"] = "التشطيب"
    ws_combined["G3"] = "الوحدة"
    for col in range(1, 8):
        ws_combined.cell(3, col).fill = cyan_fill
        ws_combined.cell(3, col).font = white_font
        ws_combined.cell(3, col).alignment = Alignment(horizontal="center")
        ws_combined.cell(3, col).border = Border(bottom=Side(style="medium"))
    
    # ملء الجداول
    total_unit_sum = unit_down
    total_finish_sum = finish_down
    total_install_sum = down_payment
    current_finish_remaining = finish_remaining
    current_unit_remaining = unit_remaining
    finish_rem = 0
    large_installments = []
    
    regular_installments = non_fixed_installments - len(annual_installments)
    regular_finish_install = round((current_finish_remaining / finish_installments) / 1000) * 1000 if regular_installments > 0 and finish_installments > 0 else 0
    if regular_finish_install > property_install * 0.6:
        regular_finish_install = round((property_install * 0.6) / 1000) * 1000
    
    row_index = 4
    for m in range(len(sorted_dates)):
        current_date = sorted_dates[m]
        current_install_num = sorted_install_nums[m]
        current_type = sorted_types[m]
        
        extra_from_date = sum(amount for j, date in enumerate(date_payments) if date == current_date)
        
        this_total_install = 0
        this_finish_install = 0
        this_unit_install = 0
        
        if current_type == "regular":
            i = current_install_num
            is_fixed_installment = i in fixed_installments
            is_extra_installment = i in extra_installments
            is_annual_installment = i in annual_installments
            
            extra_amount = extra_amounts[extra_installments.index(i)] if is_extra_installment else 0
            
            if is_fixed_installment:
                this_total_install = fixed_amount
            elif is_extra_installment:
                this_total_install = property_install + extra_amount + extra_from_date
            else:
                this_total_install = property_install + extra_from_date
                if is_annual_installment:
                    this_total_install += annual_payment
            
            if i == total_installments:
                this_total_install += property_rem
            
            if this_total_install < 0:
                this_total_install = 0
            
            ws_unit.cell(row_index, 1).value = i
            ws_unit.cell(row_index, 2).value = installment_name(i)
            ws_unit.cell(row_index, 3).value = current_date.strftime("%d-%m-%Y")
            
            ws_combined.cell(row_index, 1).value = i
            ws_combined.cell(row_index, 2).value = current_date.strftime("%d-%m-%Y")
            ws_combined.cell(row_index, 3).value = installment_name(i)
            
            if i % periods_per_year == 0:
                ws_unit.cell(row_index, 1).fill = purple_fill
                ws_unit.cell(row_index, 1).font = white_font
                ws_combined.cell(row_index, 1).fill = purple_fill
                ws_combined.cell(row_index, 1).font = white_font
                ws_combined.cell(row_index, 4).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            if i <= finish_installments and finish_total > 0 and current_finish_remaining > 0:
                ws_finish.cell(row_index, 1).value = i
                ws_finish.cell(row_index, 2).value = installment_name(i)
                ws_finish.cell(row_index, 3).value = current_date.strftime("%d-%m-%Y")
                
                if not is_extra_installment and not is_fixed_installment and not is_annual_installment:
                    this_finish_install = regular_finish_install
                    if this_finish_install > current_finish_remaining:
                        this_finish_install = round(current_finish_remaining / 1000) * 1000
                else:
                    this_finish_install = round((this_total_install * 0.6) / 1000) * 1000
                    if this_finish_install > current_finish_remaining:
                        this_finish_install = round(current_finish_remaining / 1000) * 1000
                    if this_finish_install > this_total_install * 0.8:
                        this_finish_install = round((this_total_install * 0.8) / 1000) * 1000
                
                if this_finish_install < 0:
                    this_finish_install = 0
                
                finish_rem += this_finish_install - round(this_finish_install / 1000) * 1000
                current_finish_remaining -= this_finish_install
                
                ws_finish.cell(row_index, 4).value = this_finish_install
                ws_finish.cell(row_index, 5).value = number_to_arabic_text(this_finish_install)
                total_finish_sum += this_finish_install
                
                if is_extra_installment or is_annual_installment or is_fixed_installment:
                    large_installments.append(i)
            else:
                current_unit_remaining -= this_total_install
        else:
            this_total_install = extra_from_date
            ws_unit.cell(row_index, 1).value = ""
            ws_unit.cell(row_index, 2).value = "دفعة إضافية بالتاريخ"
            ws_unit.cell(row_index, 3).value = current_date.strftime("%d-%m-%Y")
            
            ws_combined.cell(row_index, 1).value = ""
            ws_combined.cell(row_index, 2).value = current_date.strftime("%d-%m-%Y")
            ws_combined.cell(row_index, 3).value = "دفعة إضافية بالتاريخ"
            
            if current_date <= finish_end_date and finish_total > 0 and current_finish_remaining > 0:
                ws_finish.cell(row_index, 1).value = ""
                ws_finish.cell(row_index, 2).value = "دفعة إضافية بالتاريخ"
                ws_finish.cell(row_index, 3).value = current_date.strftime("%d-%m-%Y")
                
                this_finish_install = round((this_total_install * 0.6) / 1000) * 1000
                if this_finish_install > current_finish_remaining:
                    this_finish_install = round(current_finish_remaining / 1000) * 1000
                if this_finish_install > this_total_install * 0.8:
                    this_finish_install = round((this_total_install * 0.8) / 1000) * 1000
                
                if this_finish_install < 0:
                    this_finish_install = 0
                
                finish_rem += this_finish_install - round(this_finish_install / 1000) * 1000
                current_finish_remaining -= this_finish_install
                
                ws_finish.cell(row_index, 4).value = this_finish_install
                ws_finish.cell(row_index, 5).value = number_to_arabic_text(this_finish_install)
                total_finish_sum += this_finish_install
            else:
                current_unit_remaining -= this_total_install
        
        ws_combined.cell(row_index, 4).value = this_total_install
        ws_combined.cell(row_index, 5).value = number_to_arabic_text(this_total_install)
        ws_combined.cell(row_index, 6).value = this_finish_install
        ws_combined.cell(row_index, 7).value = this_total_install - this_finish_install
        
        total_unit_sum += this_total_install - this_finish_install
        total_install_sum += this_total_install
        
        ws_unit.cell(row_index, 4).value = this_total_install - this_finish_install
        ws_unit.cell(row_index, 5).value = number_to_arabic_text(ws_unit.cell(row_index, 4).value)
        
        for col in range(1, 6):
            ws_unit.cell(row_index, col).border = Border(top=Side(style="thin"), bottom=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"))
            ws_unit.cell(row_index, col).alignment = Alignment(horizontal="right")
            ws_unit.cell(row_index, col).font = Font(size=10, bold=True, name="Calibri")
        for col in range(1, 6):
            ws_finish.cell(row_index, col).border = Border(top=Side(style="thin"), bottom=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"))
            ws_finish.cell(row_index, col).alignment = Alignment(horizontal="right")
            ws_finish.cell(row_index, col).font = Font(size=10, bold=True, name="Calibri")
        for col in range(1, 8):
            ws_combined.cell(row_index, col).border = Border(top=Side(style="thin"), bottom=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"))
            ws_combined.cell(row_index, col).alignment = Alignment(horizontal="right")
            ws_combined.cell(row_index, col).font = Font(size=10, bold=True, name="Calibri")
        
        row_index += 1
    
    # إضافة دفعة الاستلام
    ws_unit.cell(row_index, 1).value = total_installments + 1
    ws_unit.cell(row_index, 2).value = "دفعة الاستلام"
    ws_unit.cell(row_index, 3).value = "دفعة الاستلام"
    ws_unit.cell(row_index, 4).value = delivery_payment
    ws_unit.cell(row_index, 5).value = number_to_arabic_text(delivery_payment)
    for col in range(1, 6):
        ws_unit.cell(row_index, col).border = Border(top=Side(style="thin"), bottom=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"))
        ws_unit.cell(row_index, col).alignment = Alignment(horizontal="right")
        ws_unit.cell(row_index, col).font = Font(size=10, bold=True, name="Calibri")
    
    ws_combined.cell(row_index, 1).value = total_installments + 1
    ws_combined.cell(row_index, 2).value = "دفعة الاستلام"
    ws_combined.cell(row_index, 3).value = installment_name(total_installments + 1)
    ws_combined.cell(row_index, 4).value = delivery_payment
    ws_combined.cell(row_index, 5).value = number_to_arabic_text(delivery_payment)
    ws_combined.cell(row_index, 7).value = delivery_payment
    for col in range(1, 8):
        ws_combined.cell(row_index, col).fill = green_fill
        ws_combined.cell(row_index, col).font = Font(size=11, bold=True, name="Calibri")
        ws_combined.cell(row_index, col).border = Border(top=Side(style="medium"), bottom=Side(style="medium"), left=Side(style="medium"), right=Side(style="medium"))
        ws_combined.cell(row_index, col).alignment = Alignment(horizontal="right")
    
    total_unit_sum += delivery_payment
    total_install_sum += delivery_payment
    
    row_index += 1
    
    # إضافة الإجمالي
    ws_unit.cell(row_index, 1).value = total_installments + 2
    ws_unit.cell(row_index, 2).value = "الإجمالي"
    ws_unit.cell(row_index, 4).value = total_unit_sum
    ws_unit.cell(row_index, 5).value = number_to_arabic_text(total_unit_sum)
    for col in range(1, 6):
        ws_unit.cell(row_index, col).fill = green_fill
        ws_unit.cell(row_index, col).font = Font(size=11, bold=True, name="Calibri")
        ws_unit.cell(row_index, col).border = Border(top=Side(style="medium"), bottom=Side(style="medium"), left=Side(style="medium"), right=Side(style="medium"))
        ws_unit.cell(row_index, col).alignment = Alignment(horizontal="right")
    
    row_index += 1
    ws_unit.cell(row_index, 2).value = "إجمالي الوحدة بدون المقدمة"
    ws_unit.cell(row_index, 4).value = total_unit_sum - unit_down
    ws_unit.cell(row_index, 5).value = number_to_arabic_text(total_unit_sum - unit_down)
    for col in range(1, 6):
        ws_unit.cell(row_index, 4).fill = blue_fill
        ws_unit.cell(row_index, 4).font = white_font
        ws_unit.cell(row_index, 4).border = Border(top=Side(style="medium"), bottom=Side(style="medium"), left=Side(style="medium"), right=Side(style="medium"))
        ws_unit.cell(row_index, 4).alignment = Alignment(horizontal="center")
    
    if finish_installments > 0:
        finish_row_index = 4 + len(sorted_dates)
        ws_finish.cell(finish_row_index, 1).value = finish_installments + 1
        ws_finish.cell(finish_row_index, 2).value = "الإجمالي"
        ws_finish.cell(finish_row_index, 4).value = total_finish_sum
        ws_finish.cell(finish_row_index, 5).value = number_to_arabic_text(total_finish_sum)
        for col in range(1, 6):
            ws_finish.cell(finish_row_index, col).fill = green_fill
            ws_finish.cell(finish_row_index, col).font = Font(size=11, bold=True, name="Calibri")
            ws_finish.cell(finish_row_index, col).border = Border(top=Side(style="medium"), bottom=Side(style="medium"), left=Side(style="medium"), right=Side(style="medium"))
            ws_finish.cell(finish_row_index, col).alignment = Alignment(horizontal="right")
    
    ws_combined.cell(row_index, 3).value = "الإجمالي"
    ws_combined.cell(row_index, 4).value = total_install_sum
    ws_combined.cell(row_index, 5).value = number_to_arabic_text(total_install_sum)
    ws_combined.cell(row_index, 6).value = total_finish_sum
    ws_combined.cell(row_index, 7).value = total_unit_sum
    for col in range(1, 8):
        ws_combined.cell(row_index, col).fill = blue_fill
        ws_combined.cell(row_index, col).font = white_font
        ws_combined.cell(row_index, col).border = Border(top=Side(style="medium"), bottom=Side(style="medium"), left=Side(style="medium"), right=Side(style="medium"))
        ws_combined.cell(row_index, col).alignment = Alignment(horizontal="center")
    
    # تنسيق الأعمدة
    for ws in [ws_unit, ws_finish]:
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 30
        for col in ["D"]:
            for cell in ws[col]:
                cell.number_format = '#,##0 "ج.م"'
    
    ws_combined.column_dimensions["A"].width = 5
    ws_combined.column_dimensions["B"].width = 12
    ws_combined.column_dimensions["C"].width = 15
    ws_combined.column_dimensions["D"].width = 12
    ws_combined.column_dimensions["E"].width = 30
    ws_combined.column_dimensions["F"].width = 12
    ws_combined.column_dimensions["G"].width = 12
    for col in ["D", "F", "G"]:
        for cell in ws_combined[col]:
            cell.number_format = '#,##0 "ج.م"'
    
    # حفظ الملف
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    st.success("تم إنشاء جدول الدفعات بنجاح!")
    st.download_button(
        label="تحميل ملف Excel",
        data=output,
        file_name=f"{unit_name} - Payment Schedule {datetime.now().strftime('%d-%m-%Y %H-%M-%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    if st.button("إعادة البدء"):
        st.session_state.step = 1
        st.session_state.data = {}
        st.session_state.extra_payment_temp = {"installment_num": 1, "amount": 0.0}
        st.rerun()
